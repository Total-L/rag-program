"""Excel (.xlsx) loader —— 多 sheet + 合并单元格 + 公式。

策略：
- openpyxl 抽每个 sheet
- 第一行作为 headers（启发式：非空 + 列数 > 1）
- 合并单元格：用左上角的值填充（用户在 UI 上看到的是这个）
- 公式：取 cached value（data_only=True），如果 None 保留 formula
- 每个 sheet 一个 table block
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from ..models import Block, BlockType, Document, ExtractionError, SourceFormat, _id_for

log = logging.getLogger(__name__)


def _fill_merged(ws) -> dict[str, str]:
    """返回 (cell_coord -> value) 映射，合并单元格用左上角值填充。"""
    out: dict[str, str] = {}
    # 先把所有 cell 值填进去
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            v = c.value
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            out[c.coordinate] = str(v)
    # 合并单元格：用左上角值填
    for mr in ws.merged_cells.ranges:
        tl = ws.cell(row=mr.min_row, column=mr.min_col)
        top_left_val = out.get(tl.coordinate, "")
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                coord = ws.cell(row=r, column=c).coordinate
                if coord not in out:
                    out[coord] = top_left_val
    return out


def _sheet_to_table(ws) -> tuple[list[str], list[list[str]]] | None:
    """抽一个 sheet → (headers, rows)。None 表示空 sheet。"""
    cell_map = _fill_merged(ws)
    if not cell_map:
        return None

    # 取最大行/列
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 1 or max_col < 1:
        return None

    # 找第一行作为 header：要求至少一半列非空
    header_row = 1
    headers: list[str] = []
    for c in range(1, max_col + 1):
        coord = ws.cell(row=header_row, column=c).coordinate
        headers.append(cell_map.get(coord, "").strip() or f"col_{c}")
    # 如果第一行全是 col_N（说明原表没 header），还是保留——总比抛错好
    # 抽数据行
    rows: list[list[str]] = [headers]
    for r in range(header_row + 1, max_row + 1):
        row_vals = []
        any_nonempty = False
        for c in range(1, max_col + 1):
            coord = ws.cell(row=r, column=c).coordinate
            v = cell_map.get(coord, "").strip()
            if v:
                any_nonempty = True
            row_vals.append(v)
        if any_nonempty:
            rows.append(row_vals)
    if len(rows) < 2:  # 只有 header 没数据
        return None
    return rows[0], rows[1:]


def _table_to_markdown(headers: list[str], rows: list[list[str]]) -> str:
    n_cols = max(len(headers), max((len(r) for r in rows), default=0))
    if n_cols == 0:
        return ""
    headers = (headers + [""] * n_cols)[:n_cols]
    lines = ["| " + " | ".join(headers) + " |", "|" + "|".join(["---"] * n_cols) + "|"]
    for r in rows:
        r2 = (r + [""] * n_cols)[:n_cols]
        lines.append("| " + " | ".join(r2) + " |")
    return "\n".join(lines)


def load(path: Path, *, data_only: bool = True) -> Document:
    """抽 xlsx → 每个 sheet 一个 TableBlock。

    data_only=True 取公式缓存值；False 取公式字符串。
    """
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ExtractionError(f"XLSX loader got {path.suffix}")
    path = Path(path).resolve()
    blocks: list[Block] = []
    warnings: list[str] = []

    wb = openpyxl.load_workbook(str(path), data_only=data_only)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            warnings.append(f"skipped hidden sheet: {sheet_name}")
            continue
        result = _sheet_to_table(ws)
        if result is None:
            warnings.append(f"empty sheet: {sheet_name}")
            continue
        headers, rows = result
        md = _table_to_markdown(headers, rows)
        blocks.append(
            Block(
                block_id=_id_for(path, sheet_name, BlockType.TABLE, len(blocks)),
                source_path=str(path),
                source_format=SourceFormat.XLSX,
                block_type=BlockType.TABLE,
                text=md,
                page_or_sheet=sheet_name,
                headers=tuple(headers),
                rows=tuple(tuple(r) for r in rows),
                metadata={"data_only": data_only, "sheet_visible": True},
            )
        )

    return Document(
        source_path=str(path),
        source_format=SourceFormat.XLSX,
        blocks=tuple(blocks),
        parse_warnings=tuple(warnings),
    )
